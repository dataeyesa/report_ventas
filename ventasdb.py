import os, re, sqlite3, datetime
from openpyxl import load_workbook

# ===== CONFIG =====
XLSX_PATH = r"D:\Desktop\APIFlaskBD\bd ventas.xlsx"
SHEET_NAME = "DATA"
HEADER_ROW = 2
DATA_START_ROW = 3
MIN_COL, MAX_COL = 1, 14       # A:N
BATCH_SIZE = 5000
SQLITE_PATH = r"D:\Desktop\APIFlaskBD\ventas.db"
TABLE_NAME = "ventas"
SAMPLE_ROWS = 20000             # muestreo para inferir tipos
# ===================

def norm_header(s, idx):
    s = ("" if s is None else str(s)).strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^\w]", "", s)
    if not s: s = f"col_{idx:02d}"
    if s[0].isdigit(): s = "_" + s
    return s[:60]

def guess_type(values):
    """
    Devuelve 'INTEGER' | 'REAL' | 'DATE' | 'TEXT'.
    - DATE: YYYY-MM-DD, YYYY/MM/DD, o excel date convertido (datetime/date)
    """
    def is_int(x):
        try:
            if isinstance(x, bool): return False
            return float(x).is_integer() and str(x).strip() != ""
        except: return False
    def is_real(x):
        try:
            float(x)
            return str(x).strip() != ""
        except: return False
    def is_date_like(x):
        if isinstance(x, (datetime.date, datetime.datetime)): return True
        xs = str(x).strip()
        for sep in ("-", "/"):
            parts = xs.split(sep)
            if len(parts)==3 and all(p.isdigit() for p in parts):
                y = int(parts[0]); m = int(parts[1]); d = int(parts[2])
                if 1<=m<=12 and 1<=d<=31 and 1900<=y<=2100:
                    return True
        return False

    seen_any = False
    ints=reals=dates=texts=0
    for v in values:
        if v is None or str(v).strip()=="":
            continue
        seen_any = True
        if is_date_like(v):
            dates+=1
        elif is_int(v):
            ints+=1
        elif is_real(v):
            reals+=1
        else:
            texts+=1
    if not seen_any:
        return "TEXT"
    # prioridad: si hay TEXT => TEXT; si hay fechas mayoritarias => DATE; luego INTEGER/REAL
    if texts>0: return "TEXT"
    if dates>0 and dates >= (ints+reals): return "TEXT" if (ints+reals)>dates else "TEXT" if dates==0 else "TEXT"  # fechas mezcladas con números => mejor TEXT
    # si no hubo texts, decide entre DATE/INTEGER/REAL:
    if dates>0 and (ints+reals)==0:
        return "DATE"
    if reals>0 and ints==0:
        return "REAL"
    if reals>0 and ints>0:
        return "REAL"
    if ints>0:
        return "INTEGER"
    return "TEXT"

def to_py_value(v, coltype):
    if v is None or str(v).strip()=="":
        return None  # NULL real
    if coltype=="INTEGER":
        try: return int(float(v))
        except: return None
    if coltype=="REAL":
        try: return float(v)
        except: return None
    if coltype=="DATE":
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip().replace("/", "-")
        try:
            y,m,d = [int(p) for p in s.split("-")]
            return f"{y:04d}-{m:02d}-{d:02d}"
        except:
            return None
    return str(v)

def main():
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(XLSX_PATH)

    print("📖 Abriendo Excel (read_only)...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe hoja {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    # Encabezados
    headers_raw = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                    min_col=MIN_COL, max_col=MAX_COL, values_only=True))
    headers = [norm_header(h, i+1) for i,h in enumerate(headers_raw)]
    print("🧾 Columnas:", headers)

    # Escaneo para tipos (muestra)
    print(f"🔎 Inferencia de tipos con muestra de {SAMPLE_ROWS} filas...")
    samples = [[] for _ in headers]
    cnt = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=True):
        for i,v in enumerate(row):
            if len(samples[i]) < SAMPLE_ROWS:
                samples[i].append(v)
        cnt += 1
        if cnt >= SAMPLE_ROWS: break

    types = [guess_type(col) for col in samples]
    # Si detectamos "DATE" mezclado con números, por seguridad fuerza TEXT:
    types = [("TEXT" if t not in ("INTEGER","REAL","DATE") else t) for t in types]
    print("📐 Tipos detectados:", dict(zip(headers, types)))

    # Reabrir hoja para leer desde el inicio de datos
    wb.close()
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    # SQLite
    if os.path.exists(SQLITE_PATH):
        os.remove(SQLITE_PATH)

    conn = sqlite3.connect(SQLITE_PATH)
    cur = conn.cursor()

    # Pragmas: velocidad y tamaño (page_size requiere VACUUM; creamos limpio y luego VACUUM)
    cur.executescript("""
        PRAGMA journal_mode=OFF;
        PRAGMA synchronous=OFF;
        PRAGMA temp_store=MEMORY;
        PRAGMA cache_size=-20000;
        PRAGMA encoding = 'UTF-8';
        PRAGMA page_size = 4096;
    """)

    # Crear tabla (sin PK definida; si tienes una columna clave, dímela y la ponemos como PK)
    cols_sql = []
    for h,t in zip(headers, types):
        if t=="DATE":  # almacenamos como TEXT 'YYYY-MM-DD' para compresión y simplicidad
            t = "TEXT"
        cols_sql.append(f'"{h}" {t}')
    cur.execute(f'DROP TABLE IF EXISTS {TABLE_NAME}')
    cur.execute(f'CREATE TABLE {TABLE_NAME} ({", ".join(cols_sql)})')

    # Insert por lotes
    placeholders = ",".join(["?"]*len(headers))
    sql = f'INSERT INTO {TABLE_NAME} VALUES ({placeholders})'

    print("⬇️ Insertando por lotes…")
    batch=[]; total=0
    row_iter = ws.iter_rows(min_row=DATA_START_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=True)
    for row in row_iter:
        vals = [to_py_value(v, types[i]) for i,v in enumerate(row)]
        batch.append(tuple(vals))
        if len(batch)>=BATCH_SIZE:
            cur.executemany(sql, batch)
            conn.commit()
            total += len(batch)
            print(f"   … {total} filas")
            batch.clear()
    if batch:
        cur.executemany(sql, batch)
        conn.commit()
        total += len(batch)
        print(f"   … {total} filas")

    # Compactar y cerrar
    print("🧹 VACUUM (compactando)…")
    cur.execute("VACUUM")   # aplica page_size y compacta
    conn.commit()
    conn.close()
    wb.close()
    print("✅ Listo:", SQLITE_PATH)

if __name__ == "__main__":
    main()
